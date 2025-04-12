# core/excel_analyzer.py
import os
import pandas as pd
import openpyxl
import re
from openpyxl.utils.exceptions import InvalidFileException

class ExcelAnalyzer:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.sheets = []
        self.has_macros = False
        self.has_formulas = False
        self.has_external_connections = False
        self.external_connections = []
        self.formulas = []
        self.error = None
        
    def analyze(self):
        """Analyze the Excel file and gather information"""
        try:
            # Check if file exists
            if not os.path.exists(self.file_path):
                self.error = f"File not found: {self.file_path}"
                return False
            
            # Check if it's an Excel file
            extension = os.path.splitext(self.file_path)[1].lower()
            if extension not in ['.xls', '.xlsx', '.xlsm']:
                self.error = f"Not an Excel file: {extension}"
                return False
                
            # Check if it has macros
            self.has_macros = extension == '.xlsm'
            
            # Load the workbook
            try:
                self.workbook = openpyxl.load_workbook(self.file_path, read_only=True, keep_vba=self.has_macros)
                self.sheets = self.workbook.sheetnames
            except InvalidFileException:
                # Try with pandas if openpyxl fails
                try:
                    excel_file = pd.ExcelFile(self.file_path)
                    self.sheets = excel_file.sheet_names
                except Exception as e:
                    self.error = f"Failed to open Excel file: {str(e)}"
                    return False
            
            # Analyze formulas
            if self.workbook:
                self._analyze_formulas()
                
            # Look for external connections
            self._analyze_external_connections()
            
            return True
        except Exception as e:
            self.error = f"Error analyzing Excel file: {str(e)}"
            return False
    
    def _analyze_formulas(self):
        """Check for formulas in the Excel workbook"""
        formula_pattern = re.compile(r'=.*\(.*\)') 
        
        for sheet_name in self.sheets:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        self.has_formulas = True
                        formula_info = {
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'formula': cell.value
                        }
                        self.formulas.append(formula_info)
                        
                        # Check for specific formula types
                        if 'VLOOKUP' in cell.value or 'HLOOKUP' in cell.value:
                            formula_info['type'] = 'lookup'
                        elif 'SUM' in cell.value or 'AVERAGE' in cell.value:
                            formula_info['type'] = 'aggregation'
                        elif 'IF' in cell.value:
                            formula_info['type'] = 'conditional'
                        else:
                            formula_info['type'] = 'other'
    
    def _analyze_external_connections(self):
        """Check for external connections in the Excel file"""
        # Note: This is a simplified version. Advanced analysis would require
        # more complex parsing of the Excel file structure.
        
        # Look for common connection patterns in formulas
        connection_patterns = [
            r'ODBC', r'ADO', r'connect', r'server', r'database', r'DSN=', 
            r'Provider=', r'Data Source='
        ]
        
        for formula in self.formulas:
            for pattern in connection_patterns:
                if re.search(pattern, formula['formula'], re.IGNORECASE):
                    self.has_external_connections = True
                    self.external_connections.append({
                        'sheet': formula['sheet'],
                        'cell': formula['cell'],
                        'formula': formula['formula'],
                        'connection_type': 'formula-based'
                    })
    
    def get_complexity_score(self):
        """Calculate a complexity score for the EUDA"""
        score = 0
        
        # Add points for multiple sheets
        if len(self.sheets) > 1:
            score += min(5, len(self.sheets))
        
        # Add points for macros
        if self.has_macros:
            score += 10
        
        # Add points for formulas
        if self.has_formulas:
            score += min(10, len(self.formulas) / 10)
        
        # Add points for external connections
        if self.has_external_connections:
            score += 15
        
        return min(100, score)
    
    def get_summary(self):
        """Return a summary of the Excel file analysis"""
        return {
            'filename': os.path.basename(self.file_path),
            'file_path': self.file_path,
            'sheets': self.sheets,
            'has_macros': self.has_macros,
            'has_formulas': self.has_formulas,
            'has_external_connections': self.has_external_connections,
            'formula_count': len(self.formulas),
            'external_connection_count': len(self.external_connections),
            'complexity_score': self.get_complexity_score()
        }